"""
topology_push_v2.py
第7课产物：企业级全网自动化配置下发（SafePusher闭环版）
核心变化：
  1. 使用 SafePusher 替代 safe_push_config，实现 备份→下发→验证→回滚 完整闭环
  2. 从 SQLite 读取设备（替代 Excel）
  3. 每个角色配置下发后自动验证，失败自动回滚
"""

from jinja2 import Template
from openpyxl import load_workbook
from network_device import NetworkDevice
from safe_pusher import SafePusher
from database import get_all_devices
from concurrent.futures import ThreadPoolExecutor, as_completed
import datetime
import os
from logger_config import setup_logger

logger = setup_logger(__name__)


# def read_topology(excel_file):
#     """
#     读取拓扑设备（Excel 版，已废弃）
#     """
#     wb = load_workbook(excel_file)
#     sheet = wb.active
#     devices = {'core': [], 'agg': [], 'access': [], 'router': [], 'server': []}
#     all_devices = []
#
#     for row in sheet.iter_rows(min_row=2, values_only=True):
#         if not row or not row[0] or not row[1]:
#             continue
#
#         device_info = {
#             'device_type': 'huawei' if row[4] == 'router' else 'huawei_telnet',
#             'host': row[1],
#             'username': row[2],
#             'password': row[3],
#             'role': row[4],
#             'dept': row[5] if len(row) > 5 else '',
#             'vlan_id': row[6] if len(row) > 6 else '',
#             'vlan_name': row[7] if len(row) > 7 else '',
#             'interface': row[8] if len(row) > 8 else 'Ethernet0/0/1',
#             'ip_address': row[9] if len(row) > 9 else '',
#             'subnet_mask': row[10] if len(row) > 10 else '',
#             'uplink': row[11] if len(row) > 11 else 'GE0/0/1',
#             'mgmt_ip': row[12] if len(row) > 12 else row[1],
#         }
#
#         if device_info['role'] in devices:
#             devices[device_info['role']].append(device_info)
#             all_devices.append(device_info)
#         else:
#             logger.warning(f"未知角色 {device_info['role']}，跳过 {device_info['host']}")
#
#     wb.close()
#     return devices, all_devices


def read_topology():
    """
    从 SQLite 读取设备并按角色分组（第6课已迁移）
    【改动】原：read_topology(excel_file) 从 new_topology.xlsx 读，返回 (devices, all_devices)
      问题：第6课已把设备台账迁到 SQLite，topology_push.py 也早已改成 get_all_devices()，
            这里 v2 又退回 Excel，等于开倒车，且又依赖 new_topology.xlsx 的表结构。
      新：改为 get_all_devices()（SQLite），返回 dict，字段与 Excel 版完全一致。
    """
    devices = {'core': [], 'agg': [], 'access': [], 'router': [], 'server': []}
    for dev in get_all_devices():
        role = dev['role']
        if role in devices:
            devices[role].append(dev)
        else:
            logger.warning(f"未知角色 {role}，跳过 {dev['host']}")
    return devices


def render_template(template_file, variables):
    with open(template_file, 'r', encoding='utf-8') as f:
        template = Template(f.read())
    config_text = template.render(**variables)
    commands = [line.strip() for line in config_text.split('\n')
                if line.strip() and not line.strip().startswith('#')]
    return commands


def config_role(device_info, template_file, desc, verify=True):
    """
    通用角色配置函数
    使用 SafePusher 实现闭环
    """
    ip = device_info['host']
    logger.info(f">>> [{device_info.get('role', 'unknown')}] 连接 {ip} ...")

    device = NetworkDevice(device_info)
    pusher = SafePusher(device)

    variables = {
        'host': ip,
        'mgmt_ip': device_info.get('mgmt_ip', ip),
    }

    # 不同角色补充不同变量
    role = device_info.get('role', '')

    if role == 'core':
        variables.update({
            'downlink_trunks': [
                {'interface': 'GE0/0/1', 'description': 'Downlink_Agg1'},
                {'interface': 'GE0/0/2', 'description': 'Downlink_Agg2'},
                {'interface': 'GE0/0/3', 'description': 'Downlink_SrvSW'},
            ],
            'uplink_ip': '10.0.0.2'
        })
    elif role == 'agg':
        dept = device_info.get('dept', '')
        if '研发' in dept or '市场' in dept:
            vlan_list = "10 20"
            downlinks = [
                {'interface': 'GE0/0/1', 'dept': '研发', 'vlan': '10'},
                {'interface': 'GE0/0/2', 'dept': '市场', 'vlan': '20'}
            ]
        else:
            vlan_list = "30 40"
            downlinks = [
                {'interface': 'GE0/0/1', 'dept': '财务', 'vlan': '30'},
                {'interface': 'GE0/0/2', 'dept': '行政', 'vlan': '40'}
            ]
        variables.update({
            'dept': dept,
            'vlan_list': vlan_list,
            'uplink': device_info.get('uplink', 'GE0/0/24'),
            'downlinks': downlinks,
            'gateway': '192.168.1.254'
        })
    elif role == 'access':
        variables.update({
            'dept': device_info.get('dept', ''),
            'vlan_id': device_info.get('vlan_id', ''),
            'uplink': device_info.get('uplink', 'GE0/0/1'),
            'gateway': '192.168.1.254'
        })
    elif role == 'server':
        variables.update({
            'uplink': device_info.get('uplink', 'GE0/0/1'),
            'interface': device_info.get('interface', 'Ethernet0/0/1'),
            'gateway': '192.168.1.254'
        })
    elif role == 'router':
        variables.update({
            'interface_ge0': 'GigabitEthernet0/0/0',
            'interface_ge1': 'GigabitEthernet0/0/1',
            'ip_ge0': '10.0.0.2',
            'mask_ge0': '255.255.255.252',
            'ip_ge1': '192.168.100.2',
            'mask_ge1': '255.255.255.0',
            'next_hop': '192.168.100.1',
            'core_ip': '10.0.0.1'
        })

    commands = render_template(template_file, variables)
    result = pusher.push(commands, description=desc, verify=verify)

    icon = "✓" if result['success'] else ("↩" if result['action'] == 'rollbacked' else "✗")
    logger.info(f"[{icon}] {ip} ({role}) {result['action'].upper()}: {result['detail']}")

    return {
        'host': ip,
        'role': role,
        'status': '成功' if result['success'] else result['action'],
        'detail': result['detail']
    }


def generate_report(results, folder):
    from openpyxl import Workbook
    report_file = f"{folder}/企业网配置报告_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "配置结果"
    ws.append(['设备IP', '角色', '状态', '详情', '时间'])

    for r in results:
        ws.append([
            r['host'], r['role'], r['status'], r.get('detail', ''),
            datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ])

    wb.save(report_file)
    logger.info(f">>> 报告已保存: {report_file}")
    return report_file


def main():
    # 【改动】原：devices_dict, all_devices = read_topology("new_topology.xlsx")
    #   问题：read_topology 已改成无参版（从 SQLite 读），只返回 devices_dict，不再返回 all_devices。
    #   新：devices_dict = read_topology()
    devices_dict = read_topology()

    tasks = []
    for role in ['router', 'core', 'agg', 'server', 'access']:
        for d in devices_dict.get(role, []):
            tasks.append((role, d))

    logger.info(f">>> 企业级全网配置下发开始（SafePusher闭环），共 {len(tasks)} 台设备")

    folder = f"report_{datetime.datetime.now().strftime('%Y%m%d')}"
    os.makedirs(folder, exist_ok=True)

    results = []

    # 第7课：串行下发（eNSP保守策略），但每个设备都有验证+回滚保护
    for role, dev in tasks:
        template_map = {
            'core': 'templates/core_config.j2',
            'agg': 'templates/agg_config.j2',
            'access': 'templates/access_config.j2',
            'server': 'templates/server_config.j2',
            'router': 'templates/router_config.j2'
        }

        desc_map = {
            'core': '核心交换机集中式网关配置',
            'agg': f"汇聚交换机({dev.get('dept', '')})配置",
            'access': f"接入交换机({dev.get('dept', '')})配置",
            'server': '服务器区交换机配置',
            'router': '出口路由器配置'
        }

        result = config_role(dev, template_map[role], desc_map[role], verify=True)
        results.append(result)

    generate_report(results, folder)

    # 统计
    success = sum(1 for r in results if r['status'] == '成功')
    rollbacked = sum(1 for r in results if r['status'] == 'rollbacked')
    failed = sum(1 for r in results if r['status'] not in ('成功', 'rollbacked'))

    logger.info(f"\n>>> 全网配置下发完成！")
    logger.info(f">>> 统计: 成功 {success} 台, 回滚 {rollbacked} 台, 失败 {failed} 台")
    logger.info(f">>> 请运行 topology_verify.py 验证连通性")


if __name__ == "__main__":
    main()
