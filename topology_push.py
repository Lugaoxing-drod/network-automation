"""
topology_push.py
企业级全网自动化配置下发（最终版）
"""

from jinja2 import Template
from openpyxl import load_workbook
from network_device import NetworkDevice
from concurrent.futures import ThreadPoolExecutor, as_completed
import datetime
import os
from logger_config import setup_logger

logger = setup_logger(__name__)


def read_topology(excel_file):
    wb = load_workbook(excel_file)
    sheet = wb.active
    devices = {'core': [], 'agg': [], 'access': [], 'router': [], 'server': []}
    all_devices = []
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        
        device_info = {
            # 【改动】按角色切换连接方式：
            #   原：'device_type': row[0],   ← 直接读 Excel 第1列（值全是 'huawei'，即 SSH）
            #   新：老交换机(S5700/S3700 V200R001) 的 SSH 服务器不兼容现代客户端，
            #       认证后主动断开，故 8 台交换机改用 telnet；路由器(AR2220) SSH 正常，保持 huawei。
            'device_type': 'huawei' if row[4] == 'router' else 'huawei_telnet',
            'host': row[1],
            'username': row[2],
            'password': row[3],
            'role': row[4],
            'dept': row[5] if len(row) > 5 else '',
            'vlan_id': row[6] if len(row) > 6 else '',
            'vlan_name': row[7] if len(row) > 7 else '',
            'interface': row[8] if len(row) > 8 else 'Ethernet0/0/1',
            'ip_address': row[9] if len(row) > 9 else '',
            'subnet_mask': row[10] if len(row) > 10 else '',
            'uplink': row[11] if len(row) > 11 else 'GE0/0/1',
            'mgmt_ip': row[12] if len(row) > 12 else row[1],
            'raw_row': row
        }
        
        if device_info['role'] in devices:
            devices[device_info['role']].append(device_info)
            all_devices.append(device_info)
        else:
            logger.warning(f"未知角色 {device_info['role']}，跳过 {device_info['host']}")
    
    wb.close()
    return devices, all_devices


def render_template(template_file, variables):
    with open(template_file, 'r', encoding='utf-8') as f:
        template = Template(f.read())
    config_text = template.render(**variables)
    commands = [line.strip() for line in config_text.split('\n')
                if line.strip() and not line.strip().startswith('#')]
                #去首尾空白、过滤空行、过滤`#`注释行
    return commands


def safe_push_config(device, commands, description=""):
    if not device.is_connected():
        logger.warning(f"    ✗ {device.host} 未连接，跳过")
        return False
    
    logger.info(f"    >>> 准备下发: {description}")
    
    # 【改动】原：backup = device.send_command("display current-configuration")
    #             logger.info(f"    ✓ 配置已备份")
    #   问题：配置只读到变量 backup，没写文件、后面也没用，等于"假备份"；
    #        真出问题想回滚时无文件可用，日志却报"已备份"，误导。
    #   新：调 device.backup(folder, timestamp) 把当前配置真正落盘到文件。
    try:
        backup_folder = f"backup_{datetime.datetime.now().strftime('%Y%m%d')}"
        os.makedirs(backup_folder, exist_ok=True)
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        if not device.backup(backup_folder, timestamp):
        # 先执行备份
        # backup_result = device.backup(backup_folder, timestamp)
        # #判断备份结果
        # if backup_result == False:
        #     logger.error(f"    ⚠ 备份失败，跳过下发")
        #     return False
        # 备份成功走到这里
        # logger.info(f"    ✓ 配置已备份")
            logger.error(f"    ⚠ 备份失败，跳过下发")
            return False
        logger.info(f"    ✓ 配置已备份")
    except Exception as e:
        logger.error(f"    ⚠ 备份失败: {e}")
        return False
    
    try:
        device.send_config(commands)
        logger.info(f"    ✓ 配置下发成功")
    except Exception as e:
        logger.error(f"    ✗ 配置下发失败: {e}")
        return False
    
    try:
        device.connection.save_config()
        logger.info(f"    ✓ 配置已保存")
    except Exception as e:
        logger.error(f"    ⚠ 保存失败: {e}")
    
    return True


def config_core(device_info):
    ip = device_info['host']
    logger.info(f">>> [核心] 连接 {ip} ...")
    
    device = NetworkDevice(device_info)
    if not device.connect():
        return {'host': ip, 'role': 'core', 'status': '失败'}
    
    variables = {
        'host': ip,
        'mgmt_ip': device_info['mgmt_ip'],
        'downlink_trunks': [
            {'interface': 'GE0/0/1', 'description': 'Downlink_Agg1'},
            {'interface': 'GE0/0/2', 'description': 'Downlink_Agg2'},
            {'interface': 'GE0/0/3', 'description': 'Downlink_SrvSW'},
        ],
        'uplink_ip': '10.0.0.2'
    }
    
    commands = render_template('templates/core_config.j2', variables)
    success = safe_push_config(device, commands, "核心交换机集中式网关配置")
    device.disconnect()
    return {'host': ip, 'role': 'core', 'status': '成功' if success else '失败'}


def config_agg(device_info):
    ip = device_info['host']
    dept = device_info['dept']
    logger.info(f">>> [汇聚] 连接 {ip} ({dept}) ...")
    
    device = NetworkDevice(device_info)
    if not device.connect():
        return {'host': ip, 'role': 'agg', 'status': '失败'}
    
    if '研发' in dept or '市场' in dept:
        vlan_list = "10 20"
        downlinks = [
            {'interface': 'GE0/0/1', 'dept': '研发', 'vlan': '10'},
            {'interface': 'GE0/0/2', 'dept': '市场', 'vlan': '20'}
        ]
    else:
        vlan_list = "30 40"
        downlinks = [
            {'interface': 'GE0/0/1', 'dept': '财务', 'vlan': '30'},
            {'interface': 'GE0/0/2', 'dept': '行政', 'vlan': '40'}
        ]
    
    variables = {
        'host': ip,
        'dept': dept,
        'mgmt_ip': device_info['mgmt_ip'],
        'vlan_list': vlan_list,
        'uplink': device_info['uplink'],
        'downlinks': downlinks,
        'gateway': '192.168.1.254'
    }
    
    commands = render_template('templates/agg_config.j2', variables)
    success = safe_push_config(device, commands, f"汇聚交换机({dept})配置")
    device.disconnect()
    return {'host': ip, 'role': 'agg', 'status': '成功' if success else '失败'}


def config_access(device_info):
    ip = device_info['host']
    dept = device_info['dept']
    logger.info(f">>> [接入] 连接 {ip} ({dept}) ...")
    
    device = NetworkDevice(device_info)
    if not device.connect():
        return {'host': ip, 'role': 'access', 'status': '失败'}
    
    variables = {
        'host': ip,
        'dept': dept,
        'vlan_id': device_info['vlan_id'],
        'uplink': device_info['uplink'],
        'mgmt_ip': device_info['mgmt_ip'],
        'gateway': '192.168.1.254'
    }
    
    commands = render_template('templates/access_config.j2', variables)
    success = safe_push_config(device, commands, f"接入交换机({dept})配置")
    device.disconnect()
    return {'host': ip, 'role': 'access', 'status': '成功' if success else '失败'}


def config_server(device_info):
    ip = device_info['host']
    logger.info(f">>> [服务器区] 连接 {ip} ...")
    
    device = NetworkDevice(device_info)
    if not device.connect():
        return {'host': ip, 'role': 'server', 'status': '失败'}
    
    variables = {
        'host': ip,
        'uplink': device_info['uplink'],
        'interface': device_info['interface'],
        'mgmt_ip': device_info['mgmt_ip'],
        'gateway': '192.168.1.254'
    }
    
    commands = render_template('templates/server_config.j2', variables)
    success = safe_push_config(device, commands, "服务器区交换机配置")
    device.disconnect()
    return {'host': ip, 'role': 'server', 'status': '成功' if success else '失败'}


def config_router(device_info):
    ip = device_info['host']
    logger.info(f">>> [出口] 连接 {ip} ...")
    
    device = NetworkDevice(device_info) 
    if not device.connect():
        return {'host': ip, 'role': 'router', 'status': '失败'}
    
    variables = {
        'host': ip,
        'interface_ge0': 'GigabitEthernet0/0/0',
        'interface_ge1': 'GigabitEthernet0/0/1',
        'ip_ge0': '10.0.0.2',
        'mask_ge0': '255.255.255.252',
        'ip_ge1': '192.168.100.2',
        'mask_ge1': '255.255.255.0',
        'next_hop': '192.168.100.1',
        'core_ip': '10.0.0.1'
    }
    
    commands = render_template('templates/router_config.j2', variables)
    success = safe_push_config(device, commands, "出口路由器配置")
    device.disconnect()
    return {'host': ip, 'role': 'router', 'status': '成功' if success else '失败'}


def generate_report(results, folder):
    from openpyxl import Workbook
    report_file = f"{folder}/企业网配置报告_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "配置结果"
    ws.append(['设备IP', '角色', '部门', '配置状态', '时间'])
    
    for r in results:
        ws.append([r['host'], r['role'], r.get('dept', ''), r['status'],
                   datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    
    wb.save(report_file)
    logger.info(f">>> 报告已保存: {report_file}")
    return report_file


def main():
    devices_dict, all_devices = read_topology("new_topology.xlsx")
    
    tasks = []
    for role in ['router', 'core', 'agg', 'server', 'access']:
        for d in devices_dict.get(role, []):
            tasks.append((role, d))
    #找到对应设备，添加到tasks列表
    logger.info(f">>> 企业级全网配置下发开始，共 {len(tasks)} 台设备")
    logger.info(f">>> 出口:{len(devices_dict['router'])} 核心:{len(devices_dict['core'])} 汇聚:{len(devices_dict['agg'])} 接入:{len(devices_dict['access'])} 服务器:{len(devices_dict['server'])}")
    
    folder = f"report_{datetime.datetime.now().strftime('%Y%m%d')}"
    os.makedirs(folder, exist_ok=True)
    #`exist_ok=True` 参数：**如果文件夹已经存在，不会抛出报错，直接静默跳过**。
    results = []
    role_map = {
        'core': config_core,
        'agg': config_agg,
        'access': config_access,
        'router': config_router,
        'server': config_server
        #分别使用对应的函数
    }

    # 【改动】原：with ThreadPoolExecutor(max_workers=5) as executor:
    #   问题：5 台并发同时向 eNSP 灌配置，抢资源导致 interface Vlanif 更慢（更容易超时）。
    #   新：改为 1，串行下发，减轻 eNSP 压力，慢命令更容易跑完。
    with ThreadPoolExecutor(max_workers=1) as executor:
        future_to_ip = {}
        for role, dev in tasks:
        #这里的dev就是上面的d，tasks.append((role, d))，d对应的就是devices_dict里的role角色对应的设备
            future = executor.submit(role_map[role], dev)
            #真正跑函数的
            future_to_ip[future] = (dev['host'], dev.get('dept', ''))
            # 把这个**future 对象直接作为字典的 key**。
            # value 存元组：(IP,部门)。
        for future in as_completed(future_to_ip):
            ip, dept = future_to_ip[future]
            try:
                result = future.result()
                #submit 提交任务的时候，就已经查完 role_map，确定要执行哪个函数了。future 对象内部已经绑定好要跑哪个函数、参数 dev。**
                results.append(result)
                icon = "✓" if result['status'] == '成功' else "✗"
                logger.info(f"[{icon}] {ip} ({result['role']}) 配置{result['status']}")
            except Exception as e:
                logger.error(f"[✗] {ip} 异常: {e}")
                results.append({'host': ip, 'role': 'unknown', 'status': f'异常: {e}'})
    
    generate_report(results, folder)
    logger.info("\n>>> 全网配置下发完成！请运行 topology_verify.py 验证连通性")


if __name__ == "__main__":
    main()