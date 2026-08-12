import datetime
import os
from openpyxl import load_workbook, Workbook
from network_device import NetworkDevice


def read_devices(excel_file):
    wb = load_workbook(excel_file)
    sheet = wb.active
    devices = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        device_type, host, username, password = row
        devices.append({
            'device_type': device_type,
            'host': host,
            'username': username,
            'password': password,
        })
    return devices


devices = read_devices("devices.xlsx")

today = datetime.datetime.now().strftime("%Y%m%d")
now = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

folder = f"report_{today}"
os.makedirs(folder, exist_ok=True)

report_file = f"{folder}/巡检报告_{now}.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "巡检结果"

headers = ['设备IP', '设备类型', '连接状态', 'CPU使用率', '内存使用率', '接口异常数', '巡检时间']
ws.append(headers)

print(f">>> 开始巡检，共 {len(devices)} 台设备")
print(f">>> 报告将保存到: {report_file}")

for dev_info in devices:
    device = NetworkDevice(dev_info)
    
    if device.connect():
        result = device.inspect()
        device.disconnect()
    else:
        result = {
            'host': dev_info['host'],
            'device_type': dev_info['device_type'],
            'status': '失败',
            'cpu': 'N/A',
            'memory': 'N/A',
            'intf_down': 'N/A',
        }
    
    ws.append([
        result['host'], result['device_type'], result['status'],
        result['cpu'], result['memory'], result['intf_down'], now,
    ])

wb.save(report_file)
print(f"\n>>> 巡检完成！报告已保存: {report_file}")