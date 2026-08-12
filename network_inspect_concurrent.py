from concurrent.futures import ThreadPoolExecutor, as_completed
from network_device import NetworkDevice
from openpyxl import load_workbook, Workbook
import datetime
import os


def read_devices(excel_file):
    """从 Excel 读取设备列表"""
    wb = load_workbook(excel_file)
    sheet = wb.active
    devices = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        device_type, host, username, password = row
        devices.append({
            'device_type': device_type,
            'host': host,
            'username': username,
            'password': password,
        })
    return devices


def inspect_single_device(dev_info):
    """
    巡检单台设备，返回字典结果
    """
    ip = dev_info['host']
    device = NetworkDevice(dev_info)
    
    # 默认失败结果
    result = {
        'host': ip,
        'device_type': dev_info['device_type'],
        'status': '失败',
        'cpu': 'N/A',
        'memory': 'N/A',
        'intf_down': 'N/A',
    }
    
    try:
        if device.connect():
            result = device.inspect()
            device.disconnect()
    except Exception as e:
        print(f"[✗] {ip} 巡检异常: {e}")
    
    return result


def main():
    devices = read_devices("devices.xlsx")
    
    today = datetime.datetime.now().strftime("%Y%m%d")
    now = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    
    # ✅ 新增：独立的 report 文件夹
    folder = f"report_{today}"
    os.makedirs(folder, exist_ok=True)
    
    report_file = f"{folder}/巡检报告_{now}.xlsx"
    
    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "巡检结果"
    headers = ['设备IP', '设备类型', '连接状态', 'CPU使用率', '内存使用率', '接口异常数', '巡检时间']
    ws.append(headers)
    
    print(f">>> 开始并发巡检，共 {len(devices)} 台设备")
    print(f">>> 报告将保存到: {report_file}")
    print(f">>> max_workers=5，报告实时生成中...\n")
    
    with ThreadPoolExecutor(max_workers=5) as executor:
        future_to_ip = {
            executor.submit(inspect_single_device, dev): dev['host']
            for dev in devices
        }
        
        for future in as_completed(future_to_ip):
            ip = future_to_ip[future]
            
            try:
                result = future.result()
                
                ws.append([
                    result['host'],
                    result['device_type'],
                    result['status'],
                    result['cpu'],
                    result['memory'],
                    result['intf_down'],
                    now,
                ])
                
                print(f"[✓] {ip} 完成 | CPU:{result['cpu']} | 内存:{result['memory']} | 接口异常:{result['intf_down']}")
                
            except Exception as e:
                print(f"[✗] {ip} 处理结果时出错: {e}")
    
    wb.save(report_file)
    print(f"\n>>> 巡检完成！报告已保存: {report_file}")


if __name__ == "__main__":
    main()