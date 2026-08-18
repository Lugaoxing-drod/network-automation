"""
migrate_to_sqlite.py
第6课：将 Excel 设备台账迁移到 SQLite 数据库
执行方式：python migrate_to_sqlite.py
只需执行一次，以后改数据直接改 database.py 的接口或用 SQL
"""

from openpyxl import load_workbook
from database import init_db, add_device
from logger_config import setup_logger

logger = setup_logger(__name__)


def migrate_from_excel(excel_file: str, sheet_type: str = "basic"):
    """
    从 Excel 读取设备并写入 SQLite
    :param excel_file: Excel 文件路径
    :param sheet_type: 'basic' 对应 devices.xlsx（第1-4课）
                       'topology' 对应 new_topology.xlsx（第5课）
    """
    logger.info(f">>> 开始从 {excel_file} 迁移数据...")

    wb = load_workbook(excel_file)
    sheet = wb.active
    count = 0
    skip = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[1]:
            continue

        # 跳过表头行（有些Excel第一行是中文表头，第二行才是数据）
        if str(row[0]).strip().lower() in ['device_type', '设备类型']:
            continue

        if sheet_type == "topology":
            # new_topology.xlsx 的列：device_type, host, username, password, role, dept,
            #                          vlan_id, vlan_name, interface, ip_address, subnet_mask, uplink, mgmt_ip
            device = {
                'device_type': 'huawei' if row[4] == 'router' else 'huawei_telnet',
                'host': row[1],
                'username': row[2],
                'password': row[3],
                'role': row[4] if len(row) > 4 else '',
                'dept': row[5] if len(row) > 5 else '',
                'vlan_id': str(row[6]) if len(row) > 6 else '',
                'vlan_name': row[7] if len(row) > 7 else '',
                'interface': row[8] if len(row) > 8 else '',
                'ip_address': row[9] if len(row) > 9 else '',
                'subnet_mask': row[10] if len(row) > 10 else '',
                'uplink': row[11] if len(row) > 11 else '',
                'mgmt_ip': row[12] if len(row) > 12 else row[1],
            }
        else:
            # devices.xlsx 的列：device_type, host, username, password, vlan_id, vlan_name, interface, ip_address, subnet_mask
            device = {
                'device_type': row[0],
                'host': row[1],
                'username': row[2],
                'password': row[3],
                'role': 'access',  # 基础台账没有role，默认access
                'dept': '',
                'vlan_id': str(row[4]) if len(row) > 4 else '',
                'vlan_name': row[5] if len(row) > 5 else '',
                'interface': row[6] if len(row) > 6 else '',
                'ip_address': row[7] if len(row) > 7 else '',
                'subnet_mask': row[8] if len(row) > 8 else '',
                'uplink': '',
                'mgmt_ip': row[1],
            }

        if add_device(device):
            count += 1
        else:
            skip += 1

    wb.close()
    logger.info(f">>> 迁移完成：新增 {count} 台，跳过 {skip} 台（已存在）")


def main():
    # 第1步：初始化数据库（建表）
    init_db()

    # 第2步：迁移基础设备台账（devices.xlsx）
    try:
        migrate_from_excel("devices.xlsx", sheet_type="basic")
    except FileNotFoundError:
        logger.warning("devices.xlsx 不存在，跳过基础台账迁移")

    # 第3步：迁移企业网拓扑台账（new_topology.xlsx）
    # 注意：如果两台Excel有重复IP，add_device会跳过（因为host有UNIQUE约束）
    try:
        migrate_from_excel("new_topology.xlsx", sheet_type="topology")
    except FileNotFoundError:
        logger.warning("new_topology.xlsx 不存在，跳过拓扑台账迁移")

    logger.info(">>> 全部迁移完成！可以用 DB Browser for SQLite 打开 network.db 查看数据。")


if __name__ == "__main__":
    main()
