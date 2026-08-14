import datetime
import os
from openpyxl import load_workbook
from network_device import NetworkDevice  # 从类文件导入
from logger_config import setup_logger

logger = setup_logger(__name__)

def read_devices(excel_file):
    """从 Excel 读取设备列表"""
    wb = load_workbook(excel_file)
    sheet = wb.active
    devices = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # 过滤空行、残缺行
        if not row or not row[0] or not row[1]:
            continue
        # 只截取前4列，后面新增任何列都不受影响
        device_type, host, username, password = row[:4]
        devices.append({
            'device_type': device_type,
            'host': host,
            'username': username,
            'password': password,
        })
    wb.close()
    return devices


# ========== 主程序 ==========
devices = read_devices("devices.xlsx")

today = datetime.datetime.now().strftime("%Y%m%d")
now = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

folder = f"backup_{today}"
os.makedirs(folder, exist_ok=True)

logger.info(f">>> 开始备份，共 {len(devices)} 台设备")
logger.info(f">>> 保存到文件夹: {folder}")

for dev_info in devices:
    # 创建设备对象
    device = NetworkDevice(dev_info)
    
    # 连接 → 备份 → 断开，一气呵成
    if device.connect():
        device.backup(folder, now)
        device.disconnect()

logger.info("\n>>> 备份全部完成！")