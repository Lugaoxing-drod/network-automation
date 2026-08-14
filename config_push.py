"""
config_push.py
业务层：配置下发（Jinja2 模板 + 安全机制）
职责：读取模板 → 渲染变量 → 下发配置
"""

from jinja2 import Template
from openpyxl import load_workbook
from network_device import NetworkDevice
import datetime
import os


def read_devices(excel_file):
    """
    从 Excel 读取设备列表（含模板变量）
    Excel 列：device_type, host, username, password, vlan_id, vlan_name, interface, ip_address, subnet_mask
    """
    wb = load_workbook(excel_file)
    sheet = wb.active
    devices = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        device_type, host, username, password, vlan_id, vlan_name, interface, ip_address, subnet_mask = row
        devices.append({
            'device_type': device_type,
            'host': host,
            'username': username,
            'password': password,
            'vlan_id': vlan_id,
            'vlan_name': vlan_name,
            'interface': interface,
            'ip_address': ip_address,
            'subnet_mask': subnet_mask,
        })
    return devices


# def create_vlan(device: NetworkDevice, vlan_id: int, vlan_name: str) -> bool:
#     """
#     业务函数：创建 VLAN
#     """
#     if not device.is_connected():
#         print(f"    ✗ {device.host} 未连接")
#         return False
    
#     commands = [
#         f'vlan {vlan_id}',
#         f'name {vlan_name}',
#     ]
#     try:
#         device.send_config(commands)
#         print(f"    ✓ {device.host} VLAN {vlan_id} 创建成功")
#         return True
#     except Exception as e:
#         print(f"    ✗ {device.host} VLAN {vlan_id} 创建失败: {e}")
#         return False


# def config_interface_ip(device: NetworkDevice, interface: str, ip: str, mask: str) -> bool:
#     """
#     业务函数：配置接口 IP
#     """
#     if not device.is_connected():
#         return False
    
#     commands = [
#         f'interface {interface}',
#         f'ip address {ip} {mask}',
#     ]
#     try:
#         device.send_config(commands)
#         print(f"    ✓ {device.host} 接口 {interface} IP 配置成功")
#         return True
#     except Exception as e:
#         print(f"    ✗ {device.host} 接口配置失败: {e}")
#         return False
 

def render_template(template_file: str, variables: dict) -> list:
    """
    读取 Jinja2 模板，渲染变量，返回命令列表
    """
    with open(template_file, 'r', encoding='utf-8') as f:
        template = Template(f.read())
    
    config_text = template.render(**variables)
    
    # 把渲染后的文本转成命令列表（去掉空行）
    commands = [line.strip() for line in config_text.split('\n') if line.strip()]
    return commands


def safe_push_config(device: NetworkDevice, commands: list, description: str = "") -> bool:
    """
    安全下发：先备份 → 下发 → 保存
    TODO: 第7课会加入"验证失败则回滚"的完整闭环
    """
    if not device.is_connected():
        print(f"    ✗ {device.host} 未连接，跳过")
        return False
    
    print(f"    >>> 准备下发: {description}")
    
    # 1. 先备份当前配置（安全机制）
    try:
        backup_config = device.send_command("display current-configuration")
        print(f"    ✓ 配置已备份（内存中）")
    except Exception as e:
        print(f"    ⚠ 备份失败: {e}，跳过下发")
        return False
    
    # 2. 下发新配置
    try:
        output = device.send_config(commands)
        print(f"    ✓ 配置下发成功")
    except Exception as e:
        print(f"    ✗ 配置下发失败: {e}")
        return False
    
    # 3. 保存配置
    try:
        device.connection.save_config()
        print(f"    ✓ 配置已保存")
    except Exception as e:
        print(f"    ⚠ 保存失败: {e}")
        return False
    
    return True


def push_vlan_from_template(device: NetworkDevice, template_file: str, variables: dict) -> bool:
    """
    业务函数：读取 VLAN 模板，渲染后安全下发
    """
    commands = render_template(template_file, variables)
    desc = f"VLAN {variables.get('vlan_id')} ({variables.get('vlan_name')})"
    return safe_push_config(device, commands, desc)


def push_ip_from_template(device: NetworkDevice, template_file: str, variables: dict) -> bool:
    """
    业务函数：读取 IP 模板，渲染后安全下发
    """
    commands = render_template(template_file, variables)
    desc = f"接口 {variables.get('interface')} IP"
    return safe_push_config(device, commands, desc)


def main():
    devices = read_devices("devices.xlsx")
    
    print(f">>> 开始配置下发，共 {len(devices)} 台设备\n")
    
    for dev_info in devices:
        ip = dev_info['host']
        print(f">>> 正在连接 {ip} ...")
        
        device = NetworkDevice(dev_info)
        if not device.connect():
            print(f"    ✗ {ip} 连接失败，跳过\n")
            continue
        
        # ========== 下发 VLAN（使用模板）==========
        vlan_vars = {
            'vlan_id': dev_info['vlan_id'],
            'vlan_name': dev_info['vlan_name'],
            'interface': dev_info['interface'],
        }
        push_vlan_from_template(device, 'templates/vlan_config.j2', vlan_vars)
        
        # ========== 下发接口 IP（使用模板）==========
        ip_vars = {
            'interface': dev_info['interface'],
            'ip_address': dev_info['ip_address'],
            'subnet_mask': dev_info['subnet_mask'],
        }
        push_ip_from_template(device, 'templates/interface_ip.j2', ip_vars)
        
        device.disconnect()
        print(f">>> {ip} 处理完成\n")
    
    print(">>> 全部设备配置下发完成")


if __name__ == "__main__":
    main()


# main()
#   │
#   ├── push_vlan_from_template(device, template_file, variables)
#   │       │
#   │       ├── render_template('templates/vlan_config.j2', vlan_vars)  ← 生成命令
#   │       │       └── 读取.j2文件 → Jinja2渲染 → 返回['vlan 10', 'name VLAN10', ...]
#   │       │
#   │       └── safe_push_config(device, commands, desc)  ← 下发命令
#   │               ├── device.send_command("display current-configuration")  ← 备份
#   │               ├── device.send_config(commands)  ← 下发（你封装的方法）
#   │               └── device.connection.save_config()  ← 保存（Netmiko原生）
#   │
#   └── push_ip_from_template(device, template_file, variables)
#           │
#           ├── render_template('templates/interface_ip.j2', ip_vars)
#           └── safe_push_config(device, commands, desc)