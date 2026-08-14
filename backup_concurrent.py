from concurrent.futures import ThreadPoolExecutor
from network_device import NetworkDevice
from openpyxl import load_workbook
import datetime
import os
import time
from logger_config import setup_logger

logger = setup_logger(__name__)

def read_devices(excel_file):
    """从 Excel 读取设备列表（和原来一样）"""
    wb = load_workbook(excel_file)
    sheet = wb.active
    devices = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        device_type, host, username, password = row
        devices.append({
            'device_type': device_type,
            'host': host,
            'username': username,
            'password': password,
        })
    return devices


def backup_single_device(dev_info, folder, now):
    """
    备份单台设备（专门给线程池调用的函数）
    每个线程独立执行这个函数
    """
    ip = dev_info['host']
    device = NetworkDevice(dev_info)
    
    try:
        if device.connect():
            success = device.backup(folder, now)
            device.disconnect()
            if success:
                logger.info(f"[✓] {ip} 备份成功")
            else:
                logger.error(f"[✗] {ip} 备份失败")
        else:
            logger.error(f"[✗] {ip} 连接失败")
    except Exception as e:
        logger.error(f"[✗] {ip} 异常: {e}")

  
def main():
    devices = read_devices("devices.xlsx")
    today = datetime.datetime.now().strftime("%Y%m%d")
    now = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    folder = f"backup_{today}"
    os.makedirs(folder, exist_ok=True)
    
    logger.info(f">>>开始并发备份，共 {len(devices)} 台设备")
    logger.info(f">>>同时连接数: 5（max_workers=5）")
    start_time = time.time()
    
    # ==================== 核心代码：线程池 ====================
    with ThreadPoolExecutor(max_workers=5) as executor:
        for dev_info in devices:
            # 把任务扔进线程池，不阻塞，继续循环
            executor.submit(backup_single_device, dev_info, folder, now)
    # with 语句结束时会自动等待所有线程完成
    # =======================================================
    
    end_time = time.time()
    logger.info(f"\n>>>全部完成！总耗时: {end_time - start_time:.2f} 秒")
    logger.info(f">>>备份保存在: {folder}")


if __name__ == "__main__":
    main()